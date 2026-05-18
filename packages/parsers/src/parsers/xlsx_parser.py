from pathlib import Path
from typing import Any

from parsers.base import (
    BaseParser,
    ExtractedSheet,
    ExtractionError,
    ExtractionResult,
    sanitize_text,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_ROWS_PER_SHEET = 10_000
MAX_SHEETS = 20


class XLSXParser(BaseParser):
    def extract(self, path: Path) -> ExtractionResult:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets: list[ExtractedSheet] = []
            warnings: list[str] = []
            total_chars = 0

            sheet_names = workbook.sheetnames
            if len(sheet_names) > MAX_SHEETS:
                warnings.append("sheets_truncated")
                sheet_names = sheet_names[:MAX_SHEETS]

            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                if worksheet.max_row and worksheet.max_row > MAX_ROWS_PER_SHEET + 1:
                    return ExtractionResult(mime_type=XLSX_MIME, error=ExtractionError.ROWS_EXCEEDED)

                extracted = _extract_sheet(worksheet)
                if extracted is None:
                    continue
                sheets.append(extracted)
                total_chars += sum(len(value) for row in extracted.rows for value in row.values())

            return ExtractionResult(
                mime_type=XLSX_MIME,
                sheets=sheets,
                total_chars=total_chars,
                warnings=warnings,
                metadata={"parser": "xlsx"},
            )
        except Exception:
            return ExtractionResult(mime_type=XLSX_MIME, error=ExtractionError.PARSE_FAILED)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return sanitize_text(str(value))


def _extract_sheet(worksheet: Any) -> ExtractedSheet | None:
    header_row_index: int | None = None
    headers: list[str] = []
    data_rows: list[tuple[int, list[str]]] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [_cell_to_text(value) for value in row]
        if header_row_index is None:
            if any(values):
                header_row_index = row_index
                headers = values
            continue
        data_rows.append((row_index, values))

    if header_row_index is None or not headers:
        return None

    records: list[dict[str, str]] = []
    first_row = data_rows[0][0] if data_rows else header_row_index + 1
    last_row = data_rows[-1][0] if data_rows else header_row_index
    for _, values in data_rows:
        padded = values + [""] * max(0, len(headers) - len(values))
        records.append(dict(zip(headers, padded, strict=False)))

    return ExtractedSheet(
        sheet_name=worksheet.title,
        headers=headers,
        rows=records,
        row_start=first_row,
        row_end=last_row,
    )
